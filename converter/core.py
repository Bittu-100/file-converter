"""
Core module for file format conversion
Supports CSV, Excel, JSON, TSV, and more
"""

import os
import csv
import json
from pathlib import Path
from typing import Optional, List, Dict, Any
import glob


class FileConverter:
    """Main class for converting between different file formats"""
    
    SUPPORTED_FORMATS = {
        'csv': 'Comma Separated Values',
        'xlsx': 'Microsoft Excel',
        'xls': 'Microsoft Excel (Legacy)',
        'json': 'JSON',
        'tsv': 'Tab Separated Values',
        'txt': 'Text File',
    }
    
    def __init__(self):
        """Initialize the file converter"""
        self.delimiter_map = {
            'csv': ',',
            'tsv': '\t',
            'txt': ',',  # Default delimiter for txt files
        }
        self.txt_delimiter = ','  # Custom delimiter for txt files
    
    def get_supported_formats(self) -> Dict[str, str]:
        """Get dictionary of supported formats"""
        return self.SUPPORTED_FORMATS.copy()
    
    def set_txt_delimiter(self, delimiter: str) -> None:
        """Set custom delimiter for txt files"""
        if len(delimiter) != 1:
            raise ValueError("Delimiter must be a single character")
        self.txt_delimiter = delimiter
        self.delimiter_map['txt'] = delimiter
    
    def get_txt_delimiter(self) -> str:
        """Get current delimiter for txt files"""
        return self.txt_delimiter
    
    def _read_csv_like(self, file_path: str, delimiter: str = ',') -> List[Dict[str, Any]]:
        """Read CSV-like files (CSV, TSV, TXT)"""
        data = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    data.append(row)
            return data
        except Exception as e:
            raise ValueError(f"Error reading file {file_path}: {str(e)}")
    
    def _read_json(self, file_path: str) -> List[Dict[str, Any]]:
        """Read JSON file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                elif isinstance(data, dict):
                    return [data]
                else:
                    raise ValueError("JSON must contain a list or dict")
            return data
        except Exception as e:
            raise ValueError(f"Error reading JSON file: {str(e)}")
    
    def _read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Read Excel file"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            data = []
            headers = []
            
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [cell for cell in row if cell is not None]
                else:
                    row_dict = {headers[i]: row[i] for i in range(len(headers))}
                    data.append(row_dict)
            
            return data
        except ImportError:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def read_file(self, file_path: str, delimiter: Optional[str] = None) -> List[Dict[str, Any]]:
        """Read a file based on its extension
        
        Args:
            file_path: Path to the file to read
            delimiter: Optional custom delimiter for txt files
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extension = file_path.suffix.lower().lstrip('.')
        
        if extension not in self.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {extension}. Supported: {', '.join(self.SUPPORTED_FORMATS.keys())}")
        
        if extension == 'csv':
            return self._read_csv_like(str(file_path), ',')
        elif extension == 'tsv':
            return self._read_csv_like(str(file_path), '\t')
        elif extension == 'txt':
            txt_delim = delimiter if delimiter else self.txt_delimiter
            return self._read_csv_like(str(file_path), txt_delim)
        elif extension == 'json':
            return self._read_json(str(file_path))
        elif extension in ['xlsx', 'xls']:
            return self._read_excel(str(file_path))
        
        return []
    
    def _write_csv(self, data: List[Dict[str, Any]], output_path: str, delimiter: str = ',') -> None:
        """Write data to CSV file"""
        if not data:
            raise ValueError("No data to write")
        
        try:
            headers = list(data[0].keys())
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(data)
        except Exception as e:
            raise ValueError(f"Error writing CSV file: {str(e)}")
    
    def _write_json(self, data: List[Dict[str, Any]], output_path: str) -> None:
        """Write data to JSON file"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            raise ValueError(f"Error writing JSON file: {str(e)}")
    
    def _write_excel(self, data: List[Dict[str, Any]], output_path: str) -> None:
        """Write data to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            
            if not data:
                raise ValueError("No data to write")
            
            headers = list(data[0].keys())
            
            # Write headers with formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Coerce non-scalar values to safe string representations
                    val = row_data.get(header, '')
                    if isinstance(val, list):
                        try:
                            val = ','.join(str(x) for x in val)
                        except Exception:
                            val = str(val)
                    if val is None:
                        val = ''
                    cell.value = val
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
        except ImportError:
            raise ImportError("openpyxl is required for Excel support. Install with: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error writing Excel file: {str(e)}")
    
    def convert(self, input_file: str, output_file: str, delimiter: Optional[str] = None, 
                 input_delimiter: Optional[str] = None) -> str:
        """
        Convert from one format to another
        
        Args:
            input_file: Path to input file
            output_file: Path to output file
            delimiter: Optional delimiter for output (for CSV/TSV/TXT)
            input_delimiter: Optional delimiter for input txt files
        
        Returns:
            Success message with output file path
        """
        # Validate input file
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Get output format
        output_ext = Path(output_file).suffix.lower().lstrip('.')
        if output_ext not in self.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported output format: {output_ext}")
        
        # Read input file
        print(f"Reading {input_path.suffix} file...")
        # Use input_delimiter if provided, otherwise use the default or stored delimiter
        if input_path.suffix.lower() == '.txt' and input_delimiter:
            data = self.read_file(input_file, input_delimiter)
        else:
            data = self.read_file(input_file)
        
        if not data:
            raise ValueError("No data read from input file")
        
        # Write output file
        print(f"Converting to {output_ext}...")
        
        try:
            if output_ext == 'csv':
                self._write_csv(data, output_file, delimiter or ',')
            elif output_ext == 'tsv':
                self._write_csv(data, output_file, '\t')
            elif output_ext == 'txt':
                # Use provided delimiter or the custom txt delimiter or default comma
                txt_out_delim = delimiter if delimiter else self.txt_delimiter
                self._write_csv(data, output_file, txt_out_delim)
            elif output_ext == 'json':
                self._write_json(data, output_file)
            elif output_ext in ['xlsx', 'xls']:
                self._write_excel(data, output_file)
        except Exception as e:
            raise ValueError(f"Conversion failed: {str(e)}")
        
        output_path = Path(output_file).resolve()
        message = f"[OK] Successfully converted to {output_file}\n  ({len(data)} records)"
        return message
    
    def merge_files(self, file1: str, file2: str, column1: str, column2: str,
                    output_base: str = None, output_format: str = 'xlsx', 
                    join_type: str = 'left') -> list:
        """
        Merge two files on specified columns with support for multiple join types

        Args:
            file1: Path to first file
            file2: Path to second file
            column1: Column name(s) in file1 to merge on (comma-separated for multi-column)
            column2: Column name(s) in file2 to merge on (comma-separated for multi-column)
            output_base: Optional base name for outputs (if None, generated from file names)
            output_format: One of 'xlsx', 'csv', 'txt', or 'both' (default: 'xlsx')
            join_type: Type of join - 'left', 'right', 'inner', 'outer' (default: 'left')

        Returns:
            List of output file paths generated
        """
        try:
            # Validate join type
            valid_joins = ('left', 'right', 'inner', 'outer')
            join_type = (join_type or 'left').lower()
            if join_type not in valid_joins:
                raise ValueError(f"join_type must be one of: {', '.join(valid_joins)}")

            # Normalize requested formats
            fmt = (output_format or 'xlsx').lower()
            if fmt not in ('xlsx', 'csv', 'txt', 'both'):
                raise ValueError("output_format must be one of: 'xlsx', 'csv', 'txt', 'both'")

            # Read both files
            print(f"Reading {file1}...")
            data1 = self.read_file(file1)

            print(f"Reading {file2}...")
            data2 = self.read_file(file2)

            if not data1:
                raise ValueError(f"No data in file1: {file1}")
            if not data2:
                raise ValueError(f"No data in file2: {file2}")

            # Handle multi-column joins
            cols1 = [c.strip() for c in column1.split(',')]
            cols2 = [c.strip() for c in column2.split(',')]
            
            if len(cols1) != len(cols2):
                raise ValueError("Number of columns must match for multi-column join")

            # Validate columns exist
            for col in cols1:
                if col not in data1[0]:
                    raise KeyError(f"Column '{col}' not found in {file1}")
            for col in cols2:
                if col not in data2[0]:
                    raise KeyError(f"Column '{col}' not found in {file2}")

            # Create composite key function
            def make_key(row, columns):
                return '|'.join(str(row.get(col, '')) for col in columns)

            # Build lookup from file2
            lookup = {}
            for row in data2:
                key = make_key(row, cols2)
                lookup[key] = row

            # Track which columns are in file2 (excluding merge columns to avoid duplicates)
            file2_columns = [col for col in list(data2[0].keys()) if col not in cols2]

            # Track which keys were matched
            matched_keys = set()

            # Perform merge based on join type
            merged_data = []

            # LEFT JOIN: all rows from file1
            if join_type in ('left', 'outer'):
                for row in data1:
                    merged_row = row.copy()
                    key = make_key(row, cols1)

                    if key in lookup:
                        matched_keys.add(key)
                        for col in file2_columns:
                            merged_row[col] = lookup[key].get(col)
                    else:
                        for col in file2_columns:
                            merged_row[col] = None

                    merged_data.append(merged_row)

            # RIGHT JOIN: all rows from file2 that match file1
            if join_type == 'right':
                for key, row2 in lookup.items():
                    # Find matching row(s) in file1
                    found = False
                    for row1 in data1:
                        if make_key(row1, cols1) == key:
                            found = True
                            merged_row = row1.copy()
                            for col in file2_columns:
                                merged_row[col] = row2.get(col)
                            merged_data.append(merged_row)
                    
                    # If no match in file1, create row with nulls from file1
                    if not found:
                        merged_row = {col: None for col in data1[0].keys()}
                        for col in file2_columns:
                            merged_row[col] = row2.get(col)
                        merged_data.append(merged_row)

            # INNER JOIN: only matching rows
            if join_type == 'inner':
                for row in data1:
                    key = make_key(row, cols1)
                    if key in lookup:
                        merged_row = row.copy()
                        for col in file2_columns:
                            merged_row[col] = lookup[key].get(col)
                        merged_data.append(merged_row)

            # OUTER JOIN: all rows from both files
            if join_type == 'outer':
                # Add unmatched rows from file2
                for key, row2 in lookup.items():
                    if key not in matched_keys:
                        merged_row = {col: None for col in data1[0].keys()}
                        for col in file2_columns:
                            merged_row[col] = row2.get(col)
                        merged_data.append(merged_row)

            # Determine output base name
            if not output_base:
                file1_base = Path(file1).stem
                file2_base = Path(file2).stem
                output_base = f"merged_{file1_base}_{file2_base}"

            outputs = []

            # Write CSV
            if fmt in ('csv', 'both'):
                csv_path = f"{output_base}.csv"
                print(f"Writing CSV to {csv_path}...")
                self._write_csv(merged_data, csv_path, ',')
                outputs.append(csv_path)

            # Write TXT (TSV)
            if fmt in ('txt', 'both'):
                txt_path = f"{output_base}.txt"
                print(f"Writing TXT (TSV) to {txt_path}...")
                self._write_csv(merged_data, txt_path, '\t')
                outputs.append(txt_path)

            # Write Excel
            if fmt in ('xlsx', 'both'):
                xlsx_path = f"{output_base}.xlsx"
                print(f"Writing Excel to {xlsx_path}...")
                self._write_excel(merged_data, xlsx_path)
                outputs.append(xlsx_path)

            print(f"\n[{join_type.upper()} JOIN] Merged {len(data1)} rows from file1 with {len(data2)} rows from file2 = {len(merged_data)} result rows")
            return outputs

        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: {e}. Please check the file paths.")
        except KeyError as e:
            raise KeyError(f"Error: Missing column {e}. Please check the column names.")
        except Exception as e:
            raise Exception(f"Unexpected error during merge: {str(e)}")

    def merge_files_with_reference(self, reference_file: str, input_files: List[str],
                                   ref_column: str, input_column: str,
                                   output_dir: str = 'merged_results',
                                   output_format: str = 'xlsx',
                                   join_type: str = 'left',
                                   file_pattern: Optional[str] = None,
                                   search_dirs: Optional[List[str]] = None) -> List[str]:
        """
        Merge multiple files with a single reference file (one-to-many merge)
        
        Each input file is merged separately with the reference file, producing one result per input file.
        
        Args:
            reference_file: Path to the reference file to merge with
            input_files: List of file paths to merge, OR None if using file_pattern/search_dirs
            ref_column: Column name in reference file to merge on (comma-separated for multi-column)
            input_column: Column name in input files to merge on (comma-separated for multi-column)
            output_dir: Directory to save results (created if doesn't exist)
            output_format: One of 'xlsx', 'csv', 'txt', or 'both'
            join_type: Type of join - 'left', 'right', 'inner', 'outer'
            file_pattern: Optional glob pattern to find files (e.g., '*.csv', '**/*.xlsx')
            search_dirs: Optional list of directories to search for files
        
        Returns:
            List of output file paths created
        """
        try:
            # Create output directory
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            # Determine which files to process
            files_to_merge = []
            
            if search_dirs:
                # Search in specified directories
                for search_dir in search_dirs:
                    search_path = Path(search_dir)
                    pattern = file_pattern or '*.csv'
                    
                    # Support recursive search with **
                    if '**' in pattern:
                        matched = list(search_path.glob(pattern))
                    else:
                        matched = list(search_path.glob(pattern))
                    
                    files_to_merge.extend([str(f) for f in matched if f.is_file()])
            
            elif file_pattern:
                # Search in current directory with pattern
                matched = glob.glob(file_pattern, recursive=True)
                files_to_merge.extend([f for f in matched if Path(f).is_file()])
            
            else:
                # Use provided input_files list
                files_to_merge = input_files or []
            
            if not files_to_merge:
                raise ValueError("No files found to merge. Check file_pattern or search_dirs.")
            
            # Verify reference file exists
            if not Path(reference_file).exists():
                raise FileNotFoundError(f"Reference file not found: {reference_file}")
            
            print(f"Found {len(files_to_merge)} files to merge with reference file: {reference_file}\n")
            
            all_outputs = []
            
            # Merge each file with reference
            for idx, input_file in enumerate(files_to_merge, 1):
                try:
                    input_name = Path(input_file).stem
                    
                    print(f"[{idx}/{len(files_to_merge)}] Merging {input_name}...")
                    
                    # Generate output base name
                    output_base = str(output_path / f"{input_name}_merged")
                    
                    # Merge input_file with reference_file
                    outputs = self.merge_files(
                        input_file,
                        reference_file,
                        input_column,
                        ref_column,
                        output_base=output_base,
                        output_format=output_format,
                        join_type=join_type
                    )
                    
                    all_outputs.extend(outputs)
                    print(f"  ✓ Created {len(outputs)} result file(s)\n")
                
                except Exception as e:
                    print(f"  ✗ Error merging {input_file}: {str(e)}\n")
                    continue
            
            print(f"\n[COMPLETE] Merged {len(files_to_merge)} files")
            print(f"Results saved to: {output_dir}/")
            print(f"Total output files: {len(all_outputs)}")
            
            return all_outputs
        
        except Exception as e:
            raise Exception(f"Error in merge_files_with_reference: {str(e)}")
    
    def get_files_from_directories(self, directories: List[str], pattern: str = '*.csv',
                                   recursive: bool = False) -> List[str]:
        """
        Find files in specified directories matching a pattern
        
        Args:
            directories: List of directory paths to search
            pattern: File pattern (e.g., '*.csv', '*.xlsx')
            recursive: If True, search recursively in subdirectories
        
        Returns:
            List of file paths found
        """
        files = []
        
        for directory in directories:
            dir_path = Path(directory)
            
            if not dir_path.exists():
                print(f"Warning: Directory not found: {directory}")
                continue
            
            search_pattern = f"**/{pattern}" if recursive else pattern
            matched = list(dir_path.glob(search_pattern))
            
            files.extend([str(f) for f in matched if f.is_file()])
        
        return sorted(files)
